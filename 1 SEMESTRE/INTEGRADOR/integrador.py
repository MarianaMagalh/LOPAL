import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Função para mapear status para cor e emoji
def interpretar_status(status):
    status = status.upper()
    if status == "2":
        return "✅", "#d4edda"  # Verde
    elif status == "1":
        return "⚠️", "#fff3cd"  # Amarelo
    elif status == "0":
        return "❌", "#f8d7da"  # Vermelho
    else:
        return "❓", "#ffffff"

# Leitura do arquivo CSV
dados = []
with open("estoque.csv", newline='', encoding='utf-8') as csvfile:
    leitor = csv.DictReader(csvfile)
    for linha in leitor:
        dados.append(linha)

# Geração da página HTML
html = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Status das Esteiras - Monitoramento de Estoque</title>
    <style>
        body { font-family: Arial; background-color: #f4f4f4; padding: 20px; }
        h1 { text-align: center; color: #333; }
        table { width: 80%; margin: auto; border-collapse: collapse; }
        th, td { padding: 12px; text-align: center; border: 1px solid #ccc; }
    </style>
</head>
<body>
<h1>Status das Esteiras - Monitoramento de Estoque</h1>
<table>
    <tr><th>Esteira</th><th>Status</th><th>Data</th></tr>
"""

for linha in dados:
    emoji, cor = interpretar_status(linha["Status"])
    html += f'<tr style="background-color:{cor}">'
    html += f'<td>{linha["Esteira"]}</td>'
    html += f'<td>{emoji} {linha["Status"]}</td>'
    html += f'<td>{linha["Data"]}</td>'
    html += "</tr>"

html += """
</table>
</body>
</html>
"""

with open("relatorio_estoque.html", "w", encoding='utf-8') as f:
    f.write(html)

# Geração do relatório Excel
wb = Workbook()
ws = wb.active
ws.title = "Status das Esteiras"

# Cabeçalhos
ws.append(["Esteira", "Status", "Data"])

# Preenchimento de linhas
for linha in dados:
    emoji, cor_hex = interpretar_status(linha["Status"])
    ws.append([linha["Esteira"], f"{emoji} {linha['Status']}", linha["Data"]])
    fill = PatternFill(start_color=cor_hex[1:], end_color=cor_hex[1:], fill_type="solid")
    for col in range(1, 4):
        ws.cell(row=ws.max_row, column=col).fill = fill

# Salva o arquivo Excel
wb.save("relatorio_estoque.xlsx")
